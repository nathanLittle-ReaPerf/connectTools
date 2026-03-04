#!/usr/bin/env python3
"""log_insights.py — Query CloudWatch Logs Insights and export results to Excel."""

from __future__ import annotations

import argparse
import datetime as dt
import re
import sys
import time
from pathlib import Path

import boto3
from botocore.config import Config
from botocore.exceptions import ClientError

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

RETRY_CONFIG   = Config(retries={"max_attempts": 5, "mode": "adaptive"})
POLL_INTERVAL  = 2   # seconds between status checks
LOG_PREFIX     = "/aws/connect/"


# ── AWS client ────────────────────────────────────────────────────────────────

def make_client(service: str, region, profile):
    session  = boto3.Session(profile_name=profile)
    resolved = region or session.region_name
    if not resolved:
        print("Error: could not determine AWS region. Pass --region explicitly.", file=sys.stderr)
        sys.exit(1)
    return session.client(service, region_name=resolved, config=RETRY_CONFIG)


# ── Log group discovery ───────────────────────────────────────────────────────

def list_connect_log_groups(logs) -> list[str]:
    groups, token = [], None
    while True:
        kwargs: dict = {"logGroupNamePrefix": LOG_PREFIX, "limit": 50}
        if token:
            kwargs["nextToken"] = token
        try:
            resp = logs.describe_log_groups(**kwargs)
        except ClientError as e:
            code = e.response["Error"]["Code"]
            msg  = e.response["Error"]["Message"]
            print(f"Error listing log groups [{code}]: {msg}", file=sys.stderr)
            sys.exit(1)
        groups.extend(g["logGroupName"] for g in resp.get("logGroups", []))
        token = resp.get("nextToken")
        if not token:
            return sorted(groups)


def resolve_log_group(logs, log_group_arg: str | None) -> str:
    if log_group_arg:
        return log_group_arg

    groups = list_connect_log_groups(logs)
    if not groups:
        print(
            f"No {LOG_PREFIX}* log groups found. Specify one with --log-group.",
            file=sys.stderr,
        )
        sys.exit(1)
    if len(groups) == 1:
        return groups[0]

    print(f"\n  Found {len(groups)} Connect log groups:\n")
    for i, g in enumerate(groups, 1):
        print(f"    {i})  {g}")
    print()
    while True:
        val = input("  Select log group (number): ").strip()
        if val.isdigit() and 1 <= int(val) <= len(groups):
            return groups[int(val) - 1]
        print(f"  Enter a number between 1 and {len(groups)}.")


# ── Time range ────────────────────────────────────────────────────────────────

_DURATION_RE = re.compile(r"^(\d+)(m|h|d|w)$", re.IGNORECASE)

def parse_duration(s: str) -> dt.timedelta:
    m = _DURATION_RE.match(s.strip())
    if not m:
        raise ValueError(f"Unrecognized duration {s!r}. Use e.g. 24h, 7d, 30m, 2w.")
    n, unit = int(m.group(1)), m.group(2).lower()
    return dt.timedelta(
        **{"m": {"minutes": n}, "h": {"hours": n}, "d": {"days": n}, "w": {"weeks": n}}[unit]
    )


def parse_dt(s: str) -> dt.datetime:
    for fmt in ("%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return dt.datetime.strptime(s, fmt).replace(tzinfo=dt.timezone.utc)
        except ValueError:
            continue
    raise ValueError(f"Unrecognized date {s!r}. Use YYYY-MM-DD or 'YYYY-MM-DD HH:MM'.")


def resolve_time_range(args) -> tuple[int, int]:
    now = dt.datetime.now(dt.timezone.utc)
    if args.last:
        try:
            delta = parse_duration(args.last)
        except ValueError as e:
            print(f"Error: {e}", file=sys.stderr)
            sys.exit(1)
        return int((now - delta).timestamp()), int(now.timestamp())

    try:
        start = parse_dt(args.start)
        end   = parse_dt(args.end) if args.end else now
    except ValueError as e:
        print(f"Error: {e}", file=sys.stderr)
        sys.exit(1)

    if start >= end:
        print("Error: --start must be before --end.", file=sys.stderr)
        sys.exit(1)

    return int(start.timestamp()), int(end.timestamp())


# ── Query execution ───────────────────────────────────────────────────────────

def start_query(logs, log_group: str, query_str: str, start_ts: int, end_ts: int, limit: int) -> str:
    try:
        resp = logs.start_query(
            logGroupName=log_group,
            startTime=start_ts,
            endTime=end_ts,
            queryString=query_str,
            limit=limit,
        )
    except ClientError as e:
        code = e.response["Error"]["Code"]
        msg  = e.response["Error"]["Message"]
        print(f"Error starting query [{code}]: {msg}", file=sys.stderr)
        sys.exit(1)
    return resp["queryId"]


def poll_query(logs, query_id: str) -> tuple[list, dict]:
    """Poll until the query completes. Returns (results, statistics)."""
    while True:
        try:
            resp = logs.get_query_results(queryId=query_id)
        except ClientError as e:
            code = e.response["Error"]["Code"]
            msg  = e.response["Error"]["Message"]
            print(f"Error polling query [{code}]: {msg}", file=sys.stderr)
            sys.exit(1)

        status = resp["status"]
        if status == "Complete":
            print()   # clear the status line
            return resp.get("results", []), resp.get("statistics", {})
        if status in ("Failed", "Cancelled", "Timeout"):
            print(f"\nQuery ended with status: {status}", file=sys.stderr)
            sys.exit(1)

        print(f"  {status}…", end="\r", flush=True)
        time.sleep(POLL_INTERVAL)


# ── Result flattening ─────────────────────────────────────────────────────────

def flatten(raw: list) -> tuple[list[str], list[list]]:
    """Convert [{field, value}] rows → (headers, rows). Drops @ptr."""
    if not raw:
        return [], []

    headers: list[str] = []
    seen: set[str]     = set()
    for row in raw:
        for item in row:
            f = item["field"]
            if f != "@ptr" and f not in seen:
                headers.append(f)
                seen.add(f)

    rows = []
    for raw_row in raw:
        lookup = {item["field"]: item["value"] for item in raw_row}
        rows.append([lookup.get(h, "") for h in headers])

    return headers, rows


# ── Excel export ──────────────────────────────────────────────────────────────

_HEADER_FILL = None
_HEADER_FONT = None

def _init_styles():
    global _HEADER_FILL, _HEADER_FONT
    _HEADER_FILL = PatternFill("solid", fgColor="1565C0")
    _HEADER_FONT = Font(bold=True, color="FFFFFF")


def export_excel(headers: list[str], rows: list[list], path: str):
    _init_styles()
    wb = openpyxl.Workbook()
    ws = wb.active

    # Header row
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Auto-fit column widths (capped at 60)
    for col_idx, h in enumerate(headers, 1):
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        max_len = max(
            len(h),
            max((len(str(row[col_idx - 1])) for row in rows), default=0),
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    wb.save(path)


# ── Argument parsing ──────────────────────────────────────────────────────────

def parse_args():
    p = argparse.ArgumentParser(
        description="Query CloudWatch Logs Insights and export results to Excel.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
examples:
  # Last 24 hours, auto-detect log group
  %(prog)s --query call_report.sql --last 24h

  # Specific date range
  %(prog)s --query call_report.sql --start 2026-03-01 --end 2026-03-02

  # Specify log group and output file
  %(prog)s --query call_report.sql --last 7d --log-group /aws/connect/myinstance --output march.xlsx

  # List available Connect log groups
  %(prog)s --list-logs
        """,
    )
    p.add_argument("--query",     metavar="FILE",     help="Logs Insights query file (.sql / .txt)")
    p.add_argument("--log-group", metavar="NAME",     help="Log group name (auto-discovers /aws/connect/ if omitted)")
    p.add_argument("--last",      metavar="DURATION", help="Relative time range: 24h, 7d, 30m, 2w …")
    p.add_argument("--start",     metavar="DATETIME", help="Start datetime: YYYY-MM-DD or 'YYYY-MM-DD HH:MM'")
    p.add_argument("--end",       metavar="DATETIME", help="End datetime (default: now)")
    p.add_argument("--limit",     type=int, default=1000, metavar="N",
                   help="Max rows returned (default: 1000, max: 10000)")
    p.add_argument("--output",    metavar="FILE",     help="Output .xlsx path (default: results_<timestamp>.xlsx)")
    p.add_argument("--list-logs", action="store_true", help="List /aws/connect/ log groups and exit")
    p.add_argument("--region",    default=None,       help="AWS region (defaults to session/CloudShell region)")
    p.add_argument("--profile",   default=None,       help="AWS named profile")
    return p.parse_args()


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    args = parse_args()

    if not HAS_OPENPYXL:
        print(
            "Error: openpyxl is required.\n"
            "  Install it with:  pip install openpyxl",
            file=sys.stderr,
        )
        sys.exit(1)

    logs = make_client("logs", args.region, args.profile)

    # ── List mode ─────────────────────────────────────────────────────────────
    if args.list_logs:
        groups = list_connect_log_groups(logs)
        if not groups:
            print(f"No {LOG_PREFIX}* log groups found.")
        else:
            print(f"\n  {len(groups)} Connect log group(s):\n")
            for g in groups:
                print(f"    {g}")
            print()
        return

    # ── Query mode validation ─────────────────────────────────────────────────
    if not args.query:
        print("Error: --query <file> is required.", file=sys.stderr)
        sys.exit(1)
    if not args.last and not args.start:
        print("Error: specify --last <duration> or --start <date>.", file=sys.stderr)
        sys.exit(1)

    query_path = Path(args.query)
    if not query_path.exists():
        print(f"Error: query file not found: {query_path}", file=sys.stderr)
        sys.exit(1)
    query_str = query_path.read_text(encoding="utf-8").strip()

    log_group          = resolve_log_group(logs, args.log_group)
    start_ts, end_ts   = resolve_time_range(args)
    start_fmt = dt.datetime.utcfromtimestamp(start_ts).strftime("%Y-%m-%d %H:%M")
    end_fmt   = dt.datetime.utcfromtimestamp(end_ts).strftime("%Y-%m-%d %H:%M")

    print(f"\n  Log group : {log_group}")
    print(f"  Time range: {start_fmt} → {end_fmt} UTC")
    print(f"  Query     : {query_path.name}")
    print(f"  Limit     : {args.limit:,}")
    print()

    query_id           = start_query(logs, log_group, query_str, start_ts, end_ts, args.limit)
    results, stats     = poll_query(logs, query_id)

    matched  = int(stats.get("recordsMatched", 0))
    scanned  = int(stats.get("recordsScanned", 0))
    print(f"  {matched:,} row(s) matched / {scanned:,} scanned")

    headers, rows = flatten(results)
    if not rows:
        print("  No results to export.")
        return

    timestamp = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path  = args.output or f"results_{timestamp}.xlsx"
    export_excel(headers, rows, out_path)
    print(f"  Exported {len(rows):,} row(s) → {out_path}\n")


if __name__ == "__main__":
    main()
